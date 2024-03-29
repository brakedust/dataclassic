"""
tables
=======

The DataTable class is a tabular data structure with behavior similar to a pandas DataFrame object.

"""

import copy
from collections import OrderedDict
from functools import partial
from itertools import chain
from string import ascii_lowercase

backend = "pyarray"
# backend = 'numpy'

if backend == "numpy":
    try:
        from numpy import append, array
        from numpy import copy as array_copy
        from numpy import delete, max, mean, min, ndarray, std, sum, transpose, var

        array_type = ndarray
    except ImportError:
        backend = "pyarray"

if backend == "pyarray":
    from dataclassic.pyarray import (
        append,
        array,
        delete,
        max,
        mean,
        min,
        pyndarray,
        std,
        sum,
        transpose,
        var,
    )

    array_type = pyndarray
    array_copy = copy.deepcopy

SIG_DIGIT_COUNT = 6


def sigdigits(x, n, format_code="g"):
    """
    Rounds a number to the specified number of significant digits
    """
    try:
        fstring = "{0:1." + str(n - 1) + format_code + "}"
        return float(fstring.format(x))
    except ValueError:
        return str(x)


def allsamelen(lst):
    """
    Determines is all elements in the list lst have the same length
    :param lst:
    :return:
    """
    firstlen = len(lst[0])
    for item in lst[1:]:
        if len(item) != firstlen:
            return False
    return True


def parse_primitive(s):
    """
    Parses a string to a python primitive type (str, int, float, bool)
    :param s:
    :return:
    """
    retval = None
    try:
        retval = int(s)
    except ValueError:
        try:
            retval = float(s)
        except ValueError:
            pass

    if retval is None:
        if s.lower() == "true":
            return True
        elif s.lower() == "false":
            return False
        else:
            return s
    else:
        return retval


def unique(seq):
    """
    Gets the unique values in seq and returns a new list with these values.
    Yes this is similar to a set, but it preserves order
    """
    seen = set()
    seen_add = seen.add
    return [x for x in seq if not (x in seen or seen_add(x))]


class DataTable(object):
    """
    Represents a tabular array of data similar to a pandas.DataFrame object
    """

    def __init__(
        self, data, columns=None, name=None, dtype=object, index=None, fillval=None
    ):
        # self.__initializing = True
        self.columns: list[str] = copy.copy(columns)
        self.data: pyndarray = None
        self.index: list = index
        self.name: str = name
        self.metadata: dict[str, object] = {}
        # self.__initializing = False
        self.row: RowView = RowView(self)
        self.column: ColumnView = ColumnView(self)

        self.__initialized = True

        if columns is not None:
            self.columns = copy.copy(columns)

        if data is not None:
            if isinstance(data, type(self)):
                # this is essentially a copy constructor.
                self.columns = copy.copy(data.columns)
                self.data = array_copy(data.data)
                self.name = data.name
                self.metadata = copy.copy(data.metadata)
            else:
                self.data = array(data)

        if self.columns is None:
            self.columns = [ascii_lowercase[i] for i in range(len(self.data[0]))]

        if self.index is None:
            index_col = "_index"
            if index_col not in self.columns:
                index_col = "index"
            if index_col in self.columns:
                self.index = self.get_column(index_col)
                self.drop(index_col)
                colindex = self.columns.index(index_col)
                self.data = delete(self.data, colindex, 1)
                self.columns.remove(index_col)
            else:
                self.index = list(range(self.data.shape[0]))

    @staticmethod
    def from_column_dict(
        data: dict[str, list], name: str = None, dtype: type = object
    ) -> "DataTable":
        column_names = list(data.keys())

        arr = array([data[col] for col in column_names], dtype=object)
        if len(arr.shape) == 1:
            # The values in the dict were scalars so arr would be a 1d array.
            # we need to make it a 2d array
            arr = array([arr])
        else:
            arr = transpose(arr)

        return DataTable(arr, column_names, name)

    @staticmethod
    def from_row_dicts(
        data: dict[str, object],
        name: str = None,
        fillval: object = None,
        dtype: type = object,
    ) -> "DataTable":
        columns = list(data[0].keys())  # get column names from first item

        # now ensure we have all of the column names, but keep the order of
        # columns in the first item
        all_columns = set().union(*[d.keys() for d in data])
        for c in all_columns:
            if c not in columns:
                columns.append(c)

        data_list = array(
            [[row.get(c, fillval) for c in columns] for row in data], dtype=dtype
        )
        return DataTable(data_list, columns, name=name)

    def add_row(self, newrow: list) -> "DataTable":
        """
        Adds a row to the DataTable
        :param newrow:
        :return:
        """
        if len(newrow) != len(self.data[0]):
            raise ValueError("New row must be same length as the existing rows")
        else:
            # self.data.append(newrow)
            return DataTable(
                append(self.data, [newrow], axis=0),
                columns=self.columns,
                name=self.name,
            )

    def __getitem__(self, args):
        """
        Gets on or more columns
        :param args:
        :return:
        """
        # if isinstance(args, slice):
        #    return DataTable(self.data[args], columns=self.columns)
        # if isinstance(args, int):
        #    return DataTable([self.data[args]], columns=self.columns)
        # else:
        return self.get_column(args)

    def __setitem__(self, key, values):
        """
        Sets a column from
        :param key:
        :param values:
        :return:
        """
        self.set_column(key, values)

    def __delitem__(self, key):
        self.drop(key)

    def __len__(self):
        return self.data.shape[0]

    def __getattr__(self, item):
        if ("columns" in self.__dict__) and (item in self.__dict__["columns"]):
            return self.get_column(item)
        elif item in self.__dict__:
            return super(DataTable, self).__getattr__(item)

    def __setattr__(self, key, value):
        if "_DataTable__initialized" not in self.__dict__:
            # if key == '_DataTable__initializing':
            super(DataTable, self).__setattr__(key, value)
        # else:
        # if self._DataTable__initializing:
        #     super(DataTable, self).__setattr__(key, value)
        # else:
        elif key in self.__dict__:
            super(DataTable, self).__setattr__(key, value)
        else:
            self.__setitem__(key, value)

    def iter_rows(self, include_index=False):
        """
        Returns an iterator for the rows in the DataTable
        """
        return (self.row_to_dict(i, include_index) for i in range(len(self)))

    def get_column(self, cols):
        """
        Gets one or more columns from the DataTable

        :param cols: columns to get
        """
        if not isinstance(cols, (list, set, tuple)):
            cols = [cols]

        colnums = []
        colnames = []
        # print(cols)
        for col in cols:
            if isinstance(col, str):
                # if col not in self.columnns:
                #     raise KeyError(col)
                colnames.append(col)
                colnums.append(self.columns.index(col))

            elif isinstance(col, int):
                colnums.append(col)
                colnames.append(self.columns[col])
            else:
                raise TypeError("col must be a str or an int - {0}".format(col))

        # print(colnums)
        # new_data_set = self.data[:, colnums]

        if len(cols) > 1:
            new_data_set = self.data[:, colnums]
            return DataTable(new_data_set, columns=colnames)
        else:
            new_data_set = self.data[:, colnums[0]]
            # new_data_set = new_data_set.reshape((len(new_data_set), ))
            # print(new_data_set.shape)
            return new_data_set
            # try:
            #    retval = array(list(result.values())[0])
            # except ValueError:
            #    retval = list(result.values())[0]
            # return retval

    def set_column(self, column, values=None, fillval=0):
        """
        Sets the values for a given column
        :param column:
        :param values:
        :return:
        """

        if values is None:
            values = array([fillval] * len(self), dtype=object)

        if not isinstance(values, array_type):
            values = array(values)

        if column in self.columns:
            cindex = self.columns.index(column)

            self.data[:, cindex] = values  # .reshape(len(values))

            # for i in range(len(self.data)):
            #    self.data[i][cindex] = values[i]

        else:
            self.columns.append(column)

            if len(self.data) > 0:
                values = values.reshape((len(values), 1))
                self.data = append(self.data, values, axis=1)
            #    for i in range(len(self.data)):
            #        try:
            #            self.data[i].append(values[i])
            #        except IndexError:
            #            self.data[i].append(None)
            else:
                self.data = values.reshape((len(values), 1))
                # for i in range(len(values)):
                #    self.data.append([values[i]])

    def get_row(self, index):
        """
        Returns a single row from the table located at *index*

        :param int index: the index of the row to return
        """
        return self.data[index, :]

    def sort(self, column, inplace=False, ascending=True, sort_on_index=False):
        """
        Sorts the rows based on the values in a particular column
        :param column:
        :param inplace:
        :param ascending:
        :return:
        """
        if not sort_on_index:
            colindex = self.columns.index(column)
            new_index = [x for (y, x) in sorted(zip(self[column], self.index))]

            sorted_indexes = self.data[:, colindex].argsort()
            data = self.data[sorted_indexes, :]
        else:
            new_index = sorted(self.index)

            data = [x for (y, x) in sorted(zip(self.index, self.data))]

        if not ascending:
            data = data[::-1, :]
            new_index = new_index[::-1]

        if inplace:
            self.data = data
            self.index = new_index
        else:
            return DataTable(data, columns=self.columns, index=new_index)

    def where(self, mask, name=None):
        """
        Gets a subset of the DataTable rows where the mask value for that row index is true
        :param list[bool] mask:
        :param name: Name of the new DataTable
        :return:
        """
        new_data = []
        new_index = []
        for irow, m in enumerate(mask):
            if m:
                new_data.append(self.data[irow])
                new_index.append(self.index[irow])

        if len(new_data) > 0:
            return DataTable(new_data, columns=self.columns, name=name, index=new_index)
        else:
            return None

    def groupby(self, column_name, lazy_eval=False):
        """
        Groups the DataTable into a set of dicts based on common values for
        the **column_name** column
        :param column_name:
        :param bool lazy_eval: If true, then each item in the returned dict is a
            function that gets it's group, otherwise each item in the return dict is a DataTable
        :return:
        """
        group_names = set(self[column_name])

        groups = {}
        if lazy_eval:
            for gn in group_names:
                mask = self[column_name] == gn
                groups[gn] = partial(self.where, mask, name=gn)
        else:
            for gn in group_names:
                groups[gn] = self.where(self[column_name] == gn, name=gn)

        return groups

    def bin_by(self, column_name, nbins, lazy_eval=False):
        group_values = set(self[column_name])
        ming = min(group_values)
        maxg = max(group_values)
        bin_size = (maxg - ming) / nbins

        bins = [(ming + i * bin_size, ming + (i + 1) * bin_size) for i in range(nbins)]

        groups = {}
        for b in bins:
            mask = (self[column_name] >= b[0]) & (self[column_name] < b[1])
            bin_name = "{0} <= {1} < {2}".format(b[0], column_name, b[1])
            groups[bin_name] = self.where(mask, name=bin_name)

        return groups

    def intersect_column(self, other, column_name, sort=True):
        """
        Computes the values in the intersection of two data tables.  The intersection is done
        on the column specified by *column_name*.  The resulting data tables are subsets of
        the original where the specified column_name values overlap.
        """
        x_common = set(self[column_name])
        x_common = x_common.intersection(other[column_name])

        dt1 = self.where((xi in x_common for xi in self[column_name]))
        dt2 = other.where((xi in x_common for xi in other[column_name]))
        if sort:
            dt1.sort(column_name, True)
            dt2.sort(column_name, True)

        return dt1, dt2

    def drop(self, column_names):
        """
        Removes a column from the structure
        :param {list or str} column_names:
        :return:
        """
        if not isinstance(column_names, (list, tuple, set)):
            column_names = (column_names,)
        cols = [c for c in self.columns if c not in column_names]

        colindexes = [self.columns.index(cn) for cn in column_names]
        newdata = delete(self.data, colindexes, 1)

        t = DataTable(newdata, columns=cols, name=self.name, index=self.index)
        return t

    def reorder_columns(self, columns, fillval=""):
        """
        Reorders the columns in the DataTable.  If a specified column does not exist
        it is filled with *fillval*
        """
        val = []
        for col in columns:
            if col in self.columns:
                val.append(list(self[col]))
            else:
                val.append([fillval] * len(self))

        return DataTable(
            array([list(x) for x in zip(*val)]), columns=columns, index=self.index
        )

    def reset_index(self, in_place=False):
        """
        Resets the index on the table to be sequentially numbered

        :param in_place: If True then this DataTable is modified.  Otherwise a new one is returned.
        """
        if in_place:
            self.index = list(range(len(self)))
        else:
            return DataTable(
                self.data, columns=self.columns, index=list(range(len(self)))
            )

    def __repr__(self):
        if self.name not in ("", None):
            retval = "Name: {0}\n".format(self.name)
        else:
            retval = ""

        if len(self.metadata) > 0:
            for k in self.metadata:
                retval += "{0}: {1}\n".format(k, self.metadata[k])

        columns = [str(c) for c in self.columns]
        colwidth1 = [len(cn) for cn in columns]
        colwidth2 = []
        for colname in self.columns:
            colwidth2.append(max([len(str(sigdigits(x, 5))) for x in self[colname]]))

        colwidth = array([colwidth1, colwidth2]).max(axis=0)
        colwidth = [cw + 2 for cw in colwidth]

        lindex = max([len(str(ind)) for ind in self.index])

        retval += " " * lindex
        for icol, item in enumerate(columns):
            fmt = "{0:>" + str(colwidth[icol]) + "}"
            retval += fmt.format(item)
        retval += "\n"

        # retval += "-" * colwidth * len(self.columns) + '\n'
        for irow, row in enumerate(self.data):
            fmt = "{0:>" + str(lindex) + "}"
            retval += fmt.format(self.index[irow])

            for icol, item in enumerate(row):
                fmt = "{0:>" + str(colwidth[icol]) + "}"
                if isinstance(item, int):
                    retval += fmt.format(item)
                else:
                    retval += fmt.format(sigdigits(item, SIG_DIGIT_COUNT))
            retval += "\n"

        return retval

    def __str__(self):
        return self.__repr__()

    def to_csv(self, fileobj, delim=",", close_file=False):
        """
        Writes the DataTable to a csv file
        :param fileobj:
        :param delim:
        :return:
        """
        from pathlib import Path

        if isinstance(fileobj, (str, Path)):
            fileobj = open(fileobj, "w")
            close_file = True

        try:
            # if self.name is not None:
            #     fileobj.write("__TableName__{0}{1}\n".format(delim, self.name))

            # for k in self.metadata:
            #     fileobj.write(
            #         "__metadata__{0}{1}{0}{2}\n".format(delim, k, self.metadata[k])
            #     )

            cols = self.columns
            fileobj.write(delim + delim.join(cols) + "\n")

            for i, row in enumerate(self.data):
                fileobj.write(
                    str(self.index[i])
                    + delim
                    + delim.join([str(x) for x in row])
                    + "\n"
                )
        except:  # nopep8
            pass
        finally:
            if close_file:
                fileobj.close()

    if False:
        # not currently working
        def to_excel(self, filename=None, wb=None):
            """
            Writes the DataTable to an excel file
            """
            import openpyxl

            if not wb:
                wb = openpyxl.Workbook()

            sname = self.name if self.name else "data"
            ws = wb.create_sheet(self.name)
            #     wb.worksheets.append()
            #     ws = wb.add_sheet(self.name)
            # else:
            #     ws = wb.add_sheet("data")

            # write metadata
            irow = 0
            if self.name is not None:
                ws.write(irow, 0, "TableName")
                ws.write(irow, 1, self.name)

            for k in self.metadata:
                irow += 1
                ws.write(irow, 0, "Metadata")
                ws.write(irow, 1, k)
                ws.write(irow, 2, self.metadata[k])

            # write column names
            # header_style = xlwt.easyxf("font: bold on; align: wrap on, vert centre, horiz center")
            fnt = openpyxl.styles.Font()
            fnt.bold = True
            fnt.height = 240
            # fnt.name = 'Calibri'
            aln = openpyxl.styles.Alignment(horizontal="center")
            aln.horizontal = openpyxl.styles.alignment.horizontal_alignments[2]
            # header_style = openpyxl.styles.XFStyle()
            # header_style.font = fnt
            # header_style.alignment = aln

            irow += 1
            for i, col in enumerate(self.columns):
                ws.append(self.columns)

            # write the data
            for i, row in enumerate(self.data):
                irow += 1
                ws.write(irow, 0, self.index[i])
                for ix, x in enumerate(row):
                    try:
                        ws.write(irow, ix + 1, x)
                    except:  # nopep8
                        ws.write(irow, ix + 1, str(x))

            if filename:
                wb.save(filename)
            else:
                return wb

    @staticmethod
    def from_csv(fileobj, delim="\t"):
        """
        Reads a csv file and returns a DataTable
        :param fileobj:
        :param delim:
        :return:
        """

        if isinstance(fileobj, str):
            fileobj = open(fileobj, "r")

        with fileobj:
            line = fileobj.readline().strip("\n")
            name = None
            if line.startswith("__TableName__"):
                name = line.split(delim)[1]
                line = fileobj.readline().strip("\n")

            metadata = {}
            while line.startswith("__metadata__"):
                __, key, value = line.split(delim)
                metadata[key] = value
                line = fileobj.readline().strip("\n")

            columns = line.split(delim)[1:]
            rows = []
            index = []
            line = fileobj.readline().strip("\n")
            while len(line) > 0:
                row = line.split(delim)
                row = [parse_primitive(x) for x in row]
                index.append(row[0])
                row.pop(0)
                rows.append(row)
                line = fileobj.readline().strip("\n")

            fileobj.close()
            return DataTable(rows, columns=columns, name=name)
        # except:
        #    fileobj.close()

    def to_html(self, filename=None):
        html = """
<body>
  <h1> {name} </h1>
  {metadata}
  <table border=1>
{table}
  </table>
</body>"""

        header_tmpl = " " * 4 + "<thead>\n{0}\n    </thead>"
        row_tmpl = " " * 4 + "<tr>\n{0}\n    </tr>"
        item_tmpl = " " * 6 + "<td>{0}</td>"
        rows = []
        row = header_tmpl.format(
            "\n".join([item_tmpl.format(c) for c in [""] + self.columns])
        )
        rows.append(row)
        for i in range(len(self)):
            items = [item_tmpl.format(self.index[i])] + [
                item_tmpl.format(self[c][i]) for c in self.columns
            ]
            row = row_tmpl.format("\n".join(items))
            rows.append(row)
        table_text = "\n".join(rows)

        # format metadata
        meta_format = "  <h2> {0}: {1} </h2>"
        lines = [meta_format.format(k, v) for k, v in self.metadata.items()]
        md = "\n".join(lines)

        html = html.format(name=self.name, table=table_text, metadata=md)
        # print(html)
        if filename is None:
            return html
        else:
            with open(filename, "w") as f:
                f.write(html)

    @property
    def shape(self):
        return self.data.shape

    # def apply(self, func):

    def apply_func(self, func, axis, name):
        """
        Applies func to columns (if axis is 0) or to rows (if axis is 1).
        :param func:
        :param axis:
        :param name:
        :param args:
        :param kwargs:
        :return:
        """
        if axis == 0:
            res = []
            for c in self.columns:
                try:
                    m = func(self[c])
                except ValueError:
                    m = None
                res.append(m)
            return DataTable([res], columns=self.columns)

        elif axis > 0:
            res = []
            for row in self.data:
                try:
                    m = func(row)
                except ValueError:
                    m = None
                res.append([m])
            return DataTable(res, columns=[name], index=self.index)

    if backend == "numpy":

        def min(self, axis=0):
            """
            Finds the minimum value of a column (axis=0) or row (axis=1)
            :param axis:
            :return:
            """
            return self.apply_func(min, axis, "min")

        def max(self, axis=0):
            """
            Finds the maximum value of a column (axis=0) or row (axis=1)
            :param axis:
            :return:
            """
            return self.apply_func(max, axis, "max")

        def mean(self, axis=0):
            """
            Computes the mean value of a column (axis=0) or row (axis=1)
            :param axis:
            :return:
            """
            return self.apply_func(mean, axis, "mean")

        def var(self, axis=0, ddof=1):
            """
            Computes the variance of a column (axis=0) or row (axis=1).
            Computes the biased (sample) variance if ddof is 0.
            Computes the unbiased variance if ddof=1.
            :param axis:
            :param ddof: Degrees of freedom
            :return:
            """
            func = partial(var, ddof=ddof)
            return self.apply_func(func, axis, "var")

        def std(self, axis=0, ddof=1):
            """
            Computes the standard deviation of a column (axis=0) or row (axis=1).
            Computes the biased (sample) standard deviation if ddof is 0.
            Computes the unbiased standard deviation if ddof=1.
            :param axis:
            :param ddof: Degrees of freedom
            :return:
            """
            func = partial(std, ddof=ddof)
            return self.apply_func(std, axis, "var")

        def sum(self, axis=0):
            """
            Computes the mean value of a column (axis=0) or row (axis=1)
            :param axis:
            :return:
            """
            return self.apply_func(sum, axis, "sum")

    elif backend == "pyarray":

        def min(self, axis=0):
            """
            Finds the minimum value of a column (axis=0) or row (axis=1)
            :param axis:
            :return:
            """
            return self.data.min(axis)  # self.apply_func(np.min, axis, 'min')

        def max(self, axis=0):
            """
            Finds the maximum value of a column (axis=0) or row (axis=1)
            :param axis:
            :return:
            """
            return self.data.max(axis)  # self.apply_func(np.max, axis, 'max')

        def mean(self, axis=0):
            """
            Computes the mean value of a column (axis=0) or row (axis=1)
            :param axis:
            :return:
            """
            return self.data.mean(axis)  # self.apply_func(np.mean, axis, 'mean')

        def var(self, axis=0, ddof=1):
            """
            Computes the variance of a column (axis=0) or row (axis=1).
            Computes the biased (sample) variance if ddof is 0.
            Computes the unbiased variance if ddof=1.
            :param axis:
            :param ddof: Degrees of freedom
            :return:
            """
            # var = lambda x: np.var(x, ddof=ddof)
            # return self.apply_func(var, axis, 'var')
            return self.data.var(axis=axis, ddof=ddof)

        def std(self, axis=0, ddof=1):
            """
            Computes the standard deviation of a column (axis=0) or row (axis=1).
            Computes the biased (sample) standard deviation if ddof is 0.
            Computes the unbiased standard deviation if ddof=1.
            :param axis:
            :param ddof: Degrees of freedom
            :return:
            """
            # std = lambda x: np.std(x, ddof=ddof)
            # return self.apply_func(std, axis, 'var')
            return self.data.std(axis=axis, ddof=ddof)

        def sum(self, axis=0):
            """
            Computes the mean value of a column (axis=0) or row (axis=1)
            :param axis:
            :return:
            """
            # return self.apply_func(np.sum, axis, 'sum')
            return self.data.sum(axis=axis)

    def transpose(self):
        """
        Transposes a DataTable
        :return:
        """
        return DataTable(self.data.transpose(), columns=self.index, index=self.columns)

    def row_to_dict(self, row_index, include_index=False):
        """
        Turns a row into a dict
        :param row_index:
        :return:
        """
        d = OrderedDict(zip(self.columns, self.data[row_index]))
        if include_index:
            d["_index"] = self.index[row_index]
        return d

    def to_list_of_dicts(self, include_index=False):
        """
        Returns a list of dicts where each dict corresponds to a given row
        """
        return [self.row_to_dict(i, include_index) for i in range(len(self))]

    def rename_columns(self, rename_dict, inplace=False):
        """
        Renames a set of columns.  The rename_dict has keys that are old
        column names and the values are the new column names.  If inplace
        is True then this instance is modified, otherwise a new DataTable
        is returned.
        """
        cols = [rename_dict.get(c, c) for c in self.columns]
        if inplace:
            self.columns = cols
        else:
            return DataTable(self.data, columns=self.cols, index=self.index)

    @staticmethod
    def concat(*tables, fillval=None):
        """
        :param tables: DataTable objects to combine
        :param **kwargs: fillval - default = None
        """

        # fillval = kwargs.get("fillval", None)

        tables = [t for t in tables if t is not None]
        # all_cols = copy.copy(tables[0].columns)
        all_cols = unique(chain(*[t.columns for t in tables]))
        # for t in tables[1:]:
        #     for col in t.columns:
        #         if not col in all_cols:
        #             all_cols.append(col)

        big_table = tables[0].reorder_columns(all_cols, fillval=fillval)
        for t in tables[1:]:
            t = t.reorder_columns(all_cols, fillval=fillval)
            big_table.data = append(big_table.data, t.data, axis=0)
            big_table.index.extend(t.index)
            big_table.metadata.update(t.metadata)

        return big_table

    def join(table1: "DataTable", table2: "DataTable", on: str, how="inner"):
        """
        Joins two tables together based on a common key columns (**on**)
        """
        table1 = table1.sort(on)
        table2 = table2.sort(on)

        col1 = table1[on]
        # table1 = table1.sort(on)
        col2 = table2[on]
        # table2 = table1.sort(on)
        table2 = table2.drop([on])

        common = set(col1).intersection(set(col2))
        row_index1 = [i for i, c in enumerate(col1) if c in common]
        row_index2 = [i for i, c in enumerate(col2) if c in common]

        colnames1 = list(table1.columns)
        colnames2 = list(table2.columns)
        # colnames2.remove(on)

        new_column_names = colnames1 + colnames2

        rows = []
        for ri1, ri2 in zip(row_index1, row_index2):
            rows.append(table1.data._data[ri1] + table2.data._data[ri2])

        return DataTable(rows, columns=new_column_names)

    def head(self, n=5):
        newtable = DataTable(
            data=self.data[:n],
            name=self.name,
        )
        # newtable.history.appen(f".head(n={n})")
        return newtable


class RowView:
    def __init__(self, table: DataTable):
        self.table: DataTable = table

    def __getitem__(self, index):
        return DataTable(self.table.data[index, :], self.table.columns)

    def __setitem__(self, index, val):
        if index >= len(self.table.data):
            raise IndexError("Speficied index is large than the number of rows.")

        if len(val) != len(self.table.data[index]):
            raise ValueError(
                "Attempting to set row with data that has a mismatch in number of colums."
            )

        self.table.data._data[index] = val


class ColumnView:
    def __init__(self, table: DataTable) -> DataTable | pyndarray:
        self.table: DataTable = table

    def __getitem__(self, columns):
        if not isinstance(columns, (list, set, tuple)):
            columns = [columns]

        colnums = []
        colnames = []

        for col in columns:
            if isinstance(col, str):
                colnames.append(col)
                colnums.append(self.table.columns.index(col))

            elif isinstance(col, int):
                colnums.append(col)
                colnames.append(self.table.columns[col])
            else:
                raise TypeError("col must be a str or an int - {0}".format(col))

        if len(columns) > 1:
            new_data_set = self.table.data[:, colnums]
            return DataTable(new_data_set, columns=colnames)
        else:
            new_data_set = self.table.data[:, colnums[0]]

            return new_data_set

    def __setitem__(self, column, values=None, fillval=0):
        # def set_column(self, column, values=None, fillval=0):
        """
        Sets the values for a given column
        :param column:
        :param values:
        :return:
        """

        if values is None:
            values = array([fillval] * len(self), dtype=object)

        if not isinstance(values, array_type):
            values = array(values)

        if column in self.table.columns:
            cindex = self.table.columns.index(column)

            self.table.data[:, cindex] = values  # .reshape(len(values))


def sample_table():
    """
    Generates a simple sample table
    """
    a = list(range(1, 10))
    b = list(range(11, 20))
    c = list(range(21, 30))
    data = [[x, y, z] for x, y, z in zip(a, b, c)]
    cols = ["a", "b", "c"]
    return DataTable(data, columns=cols)


def sample_table2(ncol, nrow):
    import string

    from dataclassic import pyarray

    keys = string.ascii_lowercase[:ncol]
    data = pyarray.pyndarray(shape=(nrow, ncol)).random_fill()
    # data = [[random.normalvariate(0.0, 1.0) for i in range(ncol)] for j in range(nrow)]
    return DataTable(data, keys)
